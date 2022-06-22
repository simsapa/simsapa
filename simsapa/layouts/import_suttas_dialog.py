from typing import List, Optional, TypedDict
from PyQt5 import QtWidgets
from PyQt5.QtCore import pyqtSignal
from PyQt5.QtWidgets import QDialog, QFileDialog, QHBoxLayout, QLabel, QPushButton, QVBoxLayout

import re
from pathlib import Path
from bs4 import BeautifulSoup

from sqlalchemy.sql import func

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from simsapa import DbSchemaName, logger
from simsapa.app.helpers import gretil_header_to_footer

from ..app.types import AppData
from ..app.db import userdata_models as Um


class SuttaRow(TypedDict):
    html_file_path: str
    html_first_line: int
    html_last_line: int
    sutta_ref: str
    language: str
    title: str
    author_uid: str


class ImportSuttasWithSpreadsheetDialog(QDialog):

    accepted = pyqtSignal(dict) # type: ignore
    _layout: QVBoxLayout
    suttas_wb: Optional[Workbook]
    sheet_file_path: Optional[Path]

    def __init__(self, app_data: AppData, parent=None) -> None:
        super().__init__(parent)

        self.setWindowTitle("Import Suttas With Spreadsheet")

        self._app_data = app_data
        self.suttas_wb = None
        self.sheet_file_path = None

        self._ui_setup()
        self._connect_signals()

    def _ui_setup(self):
        self._layout = QVBoxLayout()

        self.open_spreadsheet_button = QPushButton("Open Spreadsheet (.xlsx) ...")
        self._layout.addWidget(self.open_spreadsheet_button)

        self.status_msg = QLabel("")
        self._layout.addWidget(self.status_msg)

        self.buttons_layout = QHBoxLayout()

        self.import_button = QPushButton("Import")
        self.import_button.setDisabled(True)
        self.buttons_layout.addWidget(self.import_button)

        self.cancel_button = QPushButton("Cancel")
        self.buttons_layout.addWidget(self.cancel_button)

        self._layout.addLayout(self.buttons_layout)

        self.setLayout(self._layout)

    def _ok_pressed(self):
        logger.info("ok_pressed()")
        self.accept()

    def open_spreadsheet(self, path: Path):
        logger.info("=== open_spreadsheet() ===")
        self.suttas_wb = load_workbook(filename=path, read_only=True)
        names: list[str] = self.suttas_wb.sheetnames
        logger.info(f"names: {names}")
        if len(names) == 0:
            logger.warn(f"Empty spreadsheet: {path}")
            return

    def import_sutta_row_to_user_db(self, sutta_row: SuttaRow):
        logger.info("=== import_sutta_row_to_user_db() ====")
        logger.info("%s" % sutta_row)
        if self.sheet_file_path is None:
            return

        html_path = self.sheet_file_path.parent.joinpath(sutta_row['html_file_path'])
        if not html_path.exists():
            logger.error(f"File doesn't exist: {html_path}")
            return

        try:
            with open(html_path, 'r', encoding='utf-8') as f:
                logger.info("open file: %s" % html_path)
                if sutta_row['html_first_line'] == 0:
                    a = 0
                else:
                    a = sutta_row['html_first_line']

                if sutta_row['html_last_line'] > 0:
                    b = sutta_row['html_last_line'] + 1
                    logger.info("Reading to lines")
                    lines = f.readlines()
                    content_html = "\n".join(lines[a:b])

                else:
                    logger.info("Reading to string")
                    html_text = f.read()

                    soup = BeautifulSoup(html_text, 'html5lib')
                    h = soup.find(name = 'body')
                    if h is not None:
                        body = h.decode_contents() # type: ignore
                    else:
                        logger.info("Missing <body>, using the entire html from %s" % html_path)
                        body = html_text

                    content_html = str(body)

        except Exception as e:
            logger.error(e)
            return

        logger.info(f"Content length: {len(content_html)}")

        content_classes = []

        if sutta_row['author_uid'] == 'gretil':
            content_classes.append('gretil')

            if sutta_row['html_last_line'] <= 0:
                content_html = gretil_header_to_footer(content_html)

        if sutta_row['language'] == 'skr':
            content_classes.append('lang-skr')

        if len(content_classes) > 0:
            s = " ".join(content_classes)
            content_html = f'<div class="{s}">' + content_html + '</div>'

        # find or create author

        author = self._app_data.db_session \
            .query(Um.Author) \
            .filter(Um.Author.uid == sutta_row['author_uid']) \
            .first()

        if author is None:
            author = Um.Author(
                uid = sutta_row['author_uid'],
                created_at = func.now(),
            )
            try:
                logger.info(f"Creating new author: {author.uid}")
                self._app_data.db_session.add(author)
                self._app_data.db_session.commit()
            except Exception as e:
                logger.error(e)

        # an4.10/en/than
        ref = sutta_row['sutta_ref'].lower().replace(' ', '')
        ref = re.sub(r'[^0-9a-z_-]', '_', ref)
        uid = f"{ref}/{sutta_row['language']}/{sutta_row['author_uid']}"

        # check for existing record and update if needed

        sutta = self._app_data.db_session \
            .query(Um.Sutta) \
            .filter(Um.Sutta.uid == uid) \
            .first()

        if sutta is None:
            # create new
            sutta = Um.Sutta(
                uid = uid,
                sutta_ref = sutta_row['sutta_ref'],
                language = sutta_row['language'],
                title = sutta_row['title'],
                content_html = content_html,
                created_at = func.now(),
            )
            sutta.author = author

            try:
                logger.info(f"Creating new sutta: {sutta.uid}")
                self._app_data.db_session.add(sutta)
                self._app_data.db_session.commit()
            except Exception as e:
                logger.error(e)

        else:
            # update existing
            sutta.sutta_ref = sutta_row['sutta_ref']
            sutta.language = sutta_row['language']
            sutta.title = sutta_row['title']
            sutta.content_html = content_html
            sutta.author = author
            sutta.updated_at = func.now()

            try:
                logger.info(f"Updating sutta: {sutta.uid}")
                self._app_data.db_session.commit()
            except Exception as e:
                logger.error(e)

        s = self._app_data.db_session \
                          .query(Um.Sutta) \
                          .filter(Um.Sutta.uid == uid) \
                          .all()

        self._app_data.search_indexed.index_suttas(DbSchemaName.UserData.value, s)

    def import_sheet(self, sheet: Worksheet):
        logger.info("=== import_sheet() ===")
        cols = next(sheet.values)
        col_to_idx = {}
        for idx, c in enumerate(cols):
            col_to_idx[c] = idx

        def to_sutta_row(row) -> SuttaRow:
            s = row[col_to_idx['html_file_path']].value
            if s is None:
                html_file_path = ""
            else:
                html_file_path = str(s)

            s = row[col_to_idx['html_first_line']].value
            if s is None:
                first_line = 0
            else:
                first_line = int(s)

            s = row[col_to_idx['html_last_line']].value
            if s is None:
                last_line = -1
            else:
                last_line = int(s)

            s = row[col_to_idx['sutta_ref']].value
            if s is None:
                sutta_ref = "unknown"
            else:
                sutta_ref = str(s)

            s = row[col_to_idx['language']].value
            if s is None:
                language = "pli"
            else:
                language = str(s).lower()

            s = row[col_to_idx['title']].value
            if s is None:
                title = ""
            else:
                title = str(s)

            s = row[col_to_idx['author_uid']].value
            if s is None:
                author_uid = "unknown"
            else:
                author_uid = str(s).lower()

            sutta_row = SuttaRow(
                html_file_path = html_file_path,
                html_first_line = first_line,
                html_last_line = last_line,
                sutta_ref = sutta_ref,
                language = language,
                title = title,
                author_uid = author_uid,
            )
            return sutta_row

        sutta_rows: List[SuttaRow] = []
        # skip first row, which is the headers row
        for row in sheet.iter_rows(min_row=2):
            sutta_rows.append(to_sutta_row(row))

        logger.info(f"sutta_rows count: {len(sutta_rows)}")
        for r in sutta_rows:
            self.import_sutta_row_to_user_db(r)

    def import_pressed(self):
        logger.info("=== import_pressed() ===")
        if self.suttas_wb is None:
            return

        names: list[str] = self.suttas_wb.sheetnames
        logger.info(f"names: {names}")
        ws: Worksheet = self.suttas_wb[names[0]]
        self.status_msg.setText("Importing...")

        # QtWidgets.qApp.processEvents()

        self.import_sheet(ws)
        self.accept()

    def unlock_import(self):
        if self.suttas_wb is not None and len(self.suttas_wb.sheetnames) > 0:
            self.import_button.setDisabled(False)
        else:
            self.import_button.setDisabled(True)

    def unlock(self):
        self.unlock_import()

    def open_file_dialog(self):
        file_path, _ = QFileDialog.getOpenFileName(
            None,
            "Open File...",
            "",
            "XLSX Files (*.xlsx)")

        if len(file_path) != 0:
            self.sheet_file_path = Path(file_path)
            self.open_spreadsheet(self.sheet_file_path)
            self.unlock()

    def cancel_pressed(self):
        self.reject()

    def _connect_signals(self):
        self.open_spreadsheet_button.clicked.connect(self.open_file_dialog)

        self.import_button.clicked.connect(self.import_pressed)

        self.cancel_button.clicked.connect(self.cancel_pressed)
